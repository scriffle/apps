#!/usr/bin/env python3
"""
Update data.json with 10 Science JIM content from the
2026 Y10 Science rotation planner spreadsheet.

Uses the Overview sheet as the authoritative source for which periods
are active each week, and the Physics lesson schedule sheet for content.
"""

import openpyxl
import json
import re
from datetime import datetime, timedelta

PLANNER = '2026 Y10 Science rotation planner.xlsx'
DATA_JSON = 'dis-place/data.json'

CLASS = '10 Science JIM'


def parse_overview_slots(wb):
    """Parse Overview sheet to get ordered teaching slots per week.

    Returns list of (date_str, period, day_text) in chronological order.
    day_text is the raw Overview cell content for that day (e.g. 'Swim', 'P3, P4', etc.)
    """
    ws = wb['Overview']

    normal_A = [
        (1, 'P3'), (1, 'P4'),  # Tuesday
        (2, 'P2'),              # Wednesday
        (4, 'P6'),              # Friday
    ]
    normal_B = [
        (1, 'P3'), (1, 'P4'),  # Tuesday
        (2, 'P3'),              # Wednesday
        (3, 'P1'), (3, 'P2'),  # Thursday
    ]

    skip_keywords = ['No classes', 'Labour Day', 'Good Friday',
                    'No Students', "Kings B'day", '10-11 exams', 'HPA',
                    'Cup Weekend', 'VCE Camp', 'Work experience',
                    'Staff conference']

    all_slots = []

    for row in ws.iter_rows(min_row=2, max_row=44, values_only=True):
        week_of = row[1]
        if week_of is None or not hasattr(week_of, 'strftime'):
            continue
        ab = str(row[2]).strip() if row[2] else ''
        if ab not in ('A', 'B'):
            continue
        lessons = int(row[8]) if row[8] is not None and isinstance(row[8], (int, float)) else 0
        if lessons == 0:
            continue

        day_info = {
            0: str(row[3]).strip() if row[3] else '',
            1: str(row[4]).strip() if row[4] else '',
            2: str(row[5]).strip() if row[5] else '',
            3: str(row[6]).strip() if row[6] else '',
            4: str(row[7]).strip() if row[7] else '',
        }

        normal = normal_A if ab == 'A' else normal_B
        week_slots = []

        for day_offset, period in normal:
            dt = day_info.get(day_offset, '')
            date_str = (week_of + timedelta(days=day_offset)).strftime('%Y-%m-%d')

            if 'Only' in dt:
                only_ps = ['P' + p for p in re.findall(r'P(\d)', dt)]
                if period in only_ps:
                    week_slots.append((date_str, period, dt))
                continue

            is_blocked = any(kw.lower() in dt.lower() for kw in skip_keywords)
            if is_blocked:
                continue

            week_slots.append((date_str, period, dt))

        # Trim to match Overview's expected lesson count
        week_str = week_of.strftime('%Y-%m-%d')
        if len(week_slots) > lessons:
            week_slots = week_slots[:lessons]
        elif len(week_slots) < lessons:
            print(f"  WARNING: Week {week_str} ({ab}): expected {lessons}, got {len(week_slots)}")

        all_slots.extend(week_slots)

    return all_slots


def parse_10sci1_content(rows):
    """Parse 10SCI1 section (no explicit dates, just lesson numbers)."""
    def s(v):
        return str(v).strip() if v is not None else ''

    content = []
    for i in range(2, 24):
        if i >= len(rows):
            break
        row = rows[i]
        num = row[0]
        if num is None:
            continue
        if isinstance(num, (int, float)):
            content.append(s(row[1]))
        elif isinstance(num, str) and any(d in num for d in ['Tue', 'Wed', 'Thu', 'Fri', 'Mon']):
            content.append(s(row[1]))
    return content


def parse_10sci_content(rows, start_idx, end_idx):
    """Parse 10SCI2/3 sections — extract just the ordered activity content.

    Returns list of activity strings (ignoring the explicit dates, since
    Overview is the authoritative source for slot dates).
    """
    def s(v):
        return str(v).strip() if v is not None else ''

    content = []
    for i in range(start_idx, min(end_idx, len(rows))):
        row = rows[i]
        num = row[0]
        if num is None or (isinstance(num, str) and '10SCI' in num):
            continue
        if not isinstance(num, (int, float)):
            continue
        content.append(s(row[2]))
    return content


def flow_content(slots, content, label):
    """Flow content through Overview slots, handling event overrides.

    Event slots (Swim, Aths) get the event label directly.
    Non-event slots consume content items sequentially.
    Events do NOT consume content positions — the content list represents
    the planned physics lessons, and events are separate.

    Returns dict of (date_str, period) -> activity.
    """
    event_keywords = {'Swim': 'Swim', 'Aths': 'Aths'}
    activity_map = {}
    content_idx = 0

    for date_str, period, day_text in slots:
        override = None
        for kw, event_label in event_keywords.items():
            if kw in day_text:
                override = event_label
                break

        if override:
            activity_map[(date_str, period)] = override
        else:
            if content_idx < len(content):
                activity_map[(date_str, period)] = content[content_idx]
                content_idx += 1
            else:
                activity_map[(date_str, period)] = 'TBD'
                print(f"  WARNING: {label} ran out of content at {date_str} {period}")

    if content_idx < len(content):
        print(f"  NOTE: {label} has {len(content) - content_idx} unused content items")

    return activity_map


def main():
    wb = openpyxl.load_workbook(PLANNER, data_only=True)
    ws_physics = wb['Physics lesson schedule']
    rows = list(ws_physics.iter_rows(min_row=1, values_only=True))

    # Step 1: Get all Overview teaching slots (authoritative for dates+periods)
    all_slots = parse_overview_slots(wb)
    print(f"Total Overview teaching slots: {len(all_slots)}")

    # Step 2: Parse content from Physics lesson schedule (ordered activity text only)
    sci1_content = parse_10sci1_content(rows)
    sci2_content = parse_10sci_content(rows, 27, 49)
    sci3_content = parse_10sci_content(rows, 52, 75)

    print(f"  10SCI1 content: {len(sci1_content)} lessons")
    print(f"  10SCI2 content: {len(sci2_content)} lessons")
    print(f"  10SCI3 content: {len(sci3_content)} lessons")

    # Step 3: Split Overview slots by rotation date range
    r1_slots = [(d, p, dt) for d, p, dt in all_slots if '2026-02-03' <= d <= '2026-03-11']
    r2_slots = [(d, p, dt) for d, p, dt in all_slots if '2026-03-12' <= d <= '2026-05-05']
    r3_slots = [(d, p, dt) for d, p, dt in all_slots if '2026-05-06' <= d <= '2026-06-09']
    post_slots = [(d, p, dt) for d, p, dt in all_slots if d > '2026-06-09']

    print(f"\n  R1 Overview slots: {len(r1_slots)}")
    print(f"  R2 Overview slots: {len(r2_slots)}")
    print(f"  R3 Overview slots: {len(r3_slots)}")
    print(f"  Post-rotation slots: {len(post_slots)}")

    # Step 4: Flow content through Overview slots for each rotation
    activity_map = {}

    r1_map = flow_content(r1_slots, sci1_content, 'R1')
    activity_map.update(r1_map)
    print(f"  R1 mapped: {len(r1_map)} activities")

    r2_map = flow_content(r2_slots, sci2_content, 'R2')
    activity_map.update(r2_map)
    print(f"  R2 mapped: {len(r2_map)} activities")

    r3_map = flow_content(r3_slots, sci3_content, 'R3')
    activity_map.update(r3_map)
    print(f"  R3 mapped: {len(r3_map)} activities")

    # Post-rotation slots: use day_text keywords or TBD
    for date_str, period, day_text in post_slots:
        key = (date_str, period)
        if 'Exam Revision' in day_text or 'exm rvn' in day_text:
            activity_map[key] = 'Exam Revision'
        elif 'GAT' in day_text:
            activity_map[key] = 'GAT'
        elif 'Swim' in day_text:
            activity_map[key] = 'Swim'
        elif 'Aths' in day_text:
            activity_map[key] = 'Aths'
        else:
            activity_map[key] = 'TBD'

    print(f"\nTotal activity assignments: {len(activity_map)}")

    # Step 5: Load and update data.json
    with open(DATA_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)

    period_order = {'P1': 1, 'P2': 2, 'P3': 3, 'P4': 4, 'P5': 5, 'P6': 6}

    used = set()
    updated = 0

    for d in data['days']:
        for ev in d['events']:
            if ev['class'] != CLASS:
                continue
            key = (d['date'], ev['period'])
            if key in activity_map:
                ev['activity'] = activity_map[key]
                used.add(key)
                updated += 1

    # Add new events
    days_by_date = {d['date']: d for d in data['days']}
    added = 0
    for key, activity in sorted(activity_map.items()):
        if key in used:
            continue
        date_str, period = key
        if date_str in days_by_date:
            day = days_by_date[date_str]
            exists = any(ev['period'] == period and ev['class'] == CLASS for ev in day['events'])
            if not exists:
                day['events'].append({
                    'period': period, 'class': CLASS, 'activity': activity
                })
                day['events'].sort(key=lambda e: period_order.get(e['period'], 99))
                added += 1
        else:
            new_day = {
                'date': date_str,
                'events': [{'period': period, 'class': CLASS, 'activity': activity}]
            }
            days_by_date[date_str] = new_day
            data['days'].append(new_day)
            added += 1

    # Remove 10 Science JIM events NOT in activity_map
    removed = 0
    for d in data['days']:
        new_events = []
        for ev in d['events']:
            if ev['class'] == CLASS:
                key = (d['date'], ev['period'])
                if key in activity_map:
                    new_events.append(ev)
                else:
                    removed += 1
            else:
                new_events.append(ev)
        d['events'] = new_events

    data['days'] = [d for d in data['days'] if d['events']]
    data['days'].sort(key=lambda d: d['date'])

    print(f"\nUpdated: {updated}, Added: {added}, Removed: {removed}")

    with open(DATA_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write('\n')

    print(f"\nWrote {DATA_JSON}")

    # Verification
    print("\n=== Verification ===")
    for label, start, end, expected in [
        ('R1', '2026-02-03', '2026-03-11', 22),
        ('R2', '2026-03-12', '2026-05-05', 23),
        ('R3', '2026-05-06', '2026-06-09', 22),
    ]:
        count = 0
        print(f"\n{label} ({start} to {end}, expected {expected}):")
        for d in data['days']:
            if d['date'] < start or d['date'] > end:
                continue
            for ev in d['events']:
                if ev['class'] == CLASS:
                    count += 1
                    print(f"  {count:>3}. {d['date']} {ev['period']}: {ev['activity']}")
        status = 'OK' if count == expected else f'MISMATCH (got {count})'
        print(f"  Total: {count} [{status}]")

    # Show post-rotation slots
    print("\nPost-rotation (Jun 9+):")
    for d in data['days']:
        if d['date'] <= '2026-06-09':
            continue
        for ev in d['events']:
            if ev['class'] == CLASS:
                print(f"  {d['date']} {ev['period']}: {ev['activity']}")


if __name__ == '__main__':
    main()
